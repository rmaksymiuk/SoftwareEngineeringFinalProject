import openpyxl

from Application import Application
from Database import Database
import xlsxwriter as xls
import pandas as pd
from openpyxl import Workbook, load_workbook


path = 'Applications/test.xlsx'
workbook = load_workbook(path)
worksheet = workbook['Status']
data = ('1123', None)
worksheet.append(data)
workbook.save(path)

database = Database()
database.loadDatabase()
database.loadPermissions()
database.loadStatus()
database.showDatabase()


# with pd.ExcelWriter(path) as writer:
#     writer.book = openpyxl.load_workbook(path)
#     df.to_excel(writer, sheet_name='Status')

# workbook = xls.Workbook(path)
# worksheet = workbook.add_worksheet()
# worksheet.write(0,0, "haha")
# workbook.close()

