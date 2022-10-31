from openpyxl import Workbook, load_workbook
import xlsxwriter

class User:
    def __init__(self,username, password):
        self.username = username
        self.password = password
        self.type = None
        self.permissions = {}


    def login(self, username, password):
        if username == self.username and password == self.password:
            print(f'login for username{self.username} was successful')
            return True
        else:
            return False


    def fillApplication(self, type, application, info):
        if type == 'ClientRequestDetails':
            # info = [ , , , , , , , , ,[],[],[],[],[],[],[]]
            workbook = xlsxwriter.Workbook(application.filename)
            worksheet = workbook.add_worksheet()
            row = 1
            listRow = 1
            col = 0
            for item in info:
                if isinstance(item, list):
                    for element in item:
                        worksheet.write(listRow,col,element)
                        listRow +=1
                else:
                    worksheet.write(row, col, item)
                col +=1
        elif type == 'FinancialRequest':
            #info = [ , [], [], , , ]
            workbook = xlsxwriter.Workbook(application.filename)
            worksheet = workbook.add_worksheet()
            row = 1
            listRow = 1
            col = 0
            for item in info:
                if isinstance(item, list):
                    for element in item:
                        worksheet.write(listRow, col, element)
                        listRow += 1
                else:
                    worksheet.write(row, col, item)
                col += 1
        elif type == 'EventPlanningRequest':
            workbook = xlsxwriter.Workbook(application.filename)
            worksheet = workbook.add_worksheet()
            row = 1
            listRow = 1
            col = 0
            for item in info:
                if isinstance(item, list):
                    for element in item:
                        worksheet.write(listRow, col, element)
                        listRow += 1
                else:
                    worksheet.write(row, col, item)
                col += 1
        elif type == 'RecruitmentRequest':
            workbook = xlsxwriter.Workbook(application.filename)
            worksheet = workbook.add_worksheet()
            row = 1
            listRow = 1
            col = 0
            for item in info:
                if isinstance(item, list):
                    for element in item:
                        worksheet.write(listRow, col, element)
                        listRow += 1
                else:
                    worksheet.write(row, col, item)
                col += 1
        return