import openpyxl
from openpyxl import Workbook, load_workbook
import os, sys, subprocess
import xlsxwriter

import pandas as pd
class Database:
    def __init__(self, userList = {}):
        self.fileList = []
        # Username : [Password, Position, Permission]
        # Permission Order:
        # [ClientRequestDetails, FinancialRequest, EventPlanningRequest, RecruitmentRequest]
        self.userList = {}
        self.path = 'Database/database.xlsx'
        self.permissions = {}
        self.status = {}

    def showDatabase(self):
        print(f'current applications: {self.fileList}')
        print(f'current users: {self.userList}')
        print(f'permissions: {self.permissions}')
        print(f'status: {self.status}')

    def addFile(self, file):
        self.fileList.append(file)

    def addUser(self, user):
        self.userList.update({user.username: [user.password, user.type, user.permissions]})


    def writeUsersToExcel(self):
        workbook = xlsxwriter.Workbook(self.path)
        worksheet = workbook.add_worksheet()
        header = ('UsernameList', 'PasswordList', 'Position', 	'Permission',	'FileList')
        worksheet.write_row(0,0,header)
        row = 1
        col = 0

        for key, value in self.userList.items():
            data = (key,value[0], value[1], value[2])
            for col in range(4):
                if isinstance(data[col], list):
                    worksheet.write(row, col, str(data[col]))
                else:
                    worksheet.write(row,col,data[col])
            row += 1
        workbook.close()
        return

    def addFileStatus(self, caseNum, status):
        workbook = load_workbook(self.path)
        worksheet = workbook['Status']
        data = (caseNum, status)
        worksheet.append(data)
        workbook.save(self.path)


    def loadDatabase(self):
        sheet = "Sheet1"
        df = pd.read_excel(self.path, sheet_name=sheet)#usecols=['UsernameList', 'PasswordList', 'FileList']
        self.fileList = df['FileList'].tolist()
        for i in range(len(df.UsernameList)):
            self.userList[df.UsernameList[i]] = df.PasswordList[i] #, df.Position[i], [y for y in df.Permission[i].split(',')]

    def loadPermissions(self):
        sheet = "Permissions"
        df = pd.read_excel(self.path, sheet_name=sheet)
        for i in range(len(df.Users)):
            self.permissions[df.Users[i]] = \
                [df.Position[i],
                 [df.ClientRequest[i], df.EventPlanning[i], df.FinancialRequest[i], df.RecruitmentRequest[i]]]

    def loadStatus(self):
        sheet = "Status"
        df = pd.read_excel(self.path, sheet_name = sheet)
        for i in range(len(df.CaseNum)):
            self.status[df.CaseNum[i]] = df.Status[i]

    def changeStatus(self,caseNum, newStatus):
        self.status[caseNum] = newStatus
        row = 1
        col = 0
        workbook = load_workbook(self.path)
        worksheet = workbook["Status"]
        for cell in worksheet['A']:
            if cell.value == caseNum:
                worksheet.cell(cell.row, 2).value = newStatus

        workbook.save(self.path)
        return
