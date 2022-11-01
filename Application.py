import csv
import os, sys, subprocess
from openpyxl import Workbook, load_workbook
from Database import Database

class Application:
    def __init__(self, filename, newFormNum):
        self.filename = filename
        self.newFormNum = newFormNum
        self.status = None


    def approveApplication(self):
        self.status = 'Approved'

    def rejectApplication(self):
        self.status = 'Rejected'

    def showStatus(self):
        print(f'the status of this applicatoin is {self.status}')


    def showApplication(self, filename):
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener, filename])

    #create applications
    def createForm(self, type):
        database = Database()
        #possible types of the forms:
        # ClientRequestDetails, FinancialRequest, EventPlanningRequest, RecruitmentRequest
        switcher = {'ClientRequestDetails': 'D3', 'FinancialRequest': 'E3',
                    'EventPlanningRequest': 'B3', 'RecruitmentRequest' : 'C3'}
        wbNumber = load_workbook(self.newFormNum)
        wsNumber = wbNumber.active
        caseNum = wsNumber[switcher.get(type)].value + 1
        wsNumber[switcher.get(type)].value = caseNum
        wbNumber.save(self.newFormNum)

        wb = load_workbook(self.filename)
        ws = wb.active
        ws['A2'].value = caseNum
        print(ws['A2'].value)
        dest_filename = 'Applications/' + type + str(caseNum) + '.xlsx'
        wb.save(dest_filename)

        #adding the file to the list with the status


        self.showApplication(dest_filename)

        return



